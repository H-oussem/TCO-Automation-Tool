import streamlit as st
import openpyxl
from io import BytesIO
import re
import difflib
import random
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter, column_index_from_string

st.title("📊 TCO Automation Tool ")

# ----------------- GLOBALS & CONSTANTS -----------------
REQUIRED_TCO_NAME_FRAGMENT = "2025.02.07_tco_c_v1.5.1".lower()
thin = Side(border_style="thin", color="000000")
used_colors = set()

# Column boundaries
BN_COL_NUM = 66  # BN column index (A=1 -> BN=66)
STYLE_MAX_COL_NUM = 76  # Extend styling through BX (A=1 -> BX=76)

# Keep the original template rows (10 & 11) untouched.
TEMPLATE_HEADER_ROW = 10
TEMPLATE_DATA_ROW = 11
# All program writing starts at row 12 (HEADER_ROW) and data at 13 (DATA_ROW)
HEADER_ROW = 12
DATA_ROW = HEADER_ROW + 1

# Column indices for convenience (these are used later)
BD_col = 56
BE_col = 57
BF_col = 58
BH_col = 60
BO_col = 67
BP_col = 68
BQ_col = 69
U_col = 21
T_col = 20  # used when writing/reading T values
BR_col = 70
BX_col = 76

SHIFT_PATTERN = re.compile(r'(\$?[A-Z]{1,3})(\$?\d+)')
NON_ALNUM_RE = re.compile(r'[^0-9a-z]+')
DOT_SPACE_RE = re.compile(r'\s+\.')
BOX_PER_LAYER_RE = re.compile(r'(\d{1,6})\s*(?:box\s*/\s*layer|box/layer|box\s*per\s*layer)', re.I)
LAYERS_RE = re.compile(r'(\d{1,6})\s*layers?', re.I)

# ----------------- HELPERS -----------------
def shift_formula_rows(formula, row_offset):
    """Shift row references inside an Excel formula by row_offset (int).
    Keeps absolute references (with $) intact.
    """
    return SHIFT_PATTERN.sub(
        lambda m: f"{m.group(1)}{int(m.group(2)) + row_offset}" if not m.group(2).startswith('$') else m.group(0),
        formula
    )

def count_nonempty_columns_after_B(sheet, max_rows_preview=50):
    max_col = sheet.max_column
    max_row_check = min(sheet.max_row + 1, max_rows_preview)
    for col in range(max_col, 2, -1):
        for row in range(1, max_row_check):
            if sheet.cell(row=row, column=col).value not in (None, ""):
                return col - 2
    return 0

def normalize_name(name):
    if not isinstance(name, str):
        return ""
    s = name.strip().lower()
    s = DOT_SPACE_RE.sub('.', s)
    s = NON_ALNUM_RE.sub(' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def extract_supplier_name_from_filename(filename):
    match = re.search(r"(.+)_Quote", filename, flags=re.IGNORECASE)
    return match.group(1).strip() if match else "Unknown"

def normalize_sheetname_for_match(s):
    return re.sub(r'[^0-9a-z]+', '', s.lower()) if isinstance(s, str) else ""

def is_logistics_sheet(sheet_name):
    if not sheet_name or not isinstance(sheet_name, str):
        return False
    s_low = sheet_name.lower()
    logist_tokens = ["logist", "logistics", "logistik"]
    surch_tokens = ["surcharg", "surcharge", "surcharges", "zuschlag", "zuschlaeg", "zuschlaege", "sonst", "misc"]
    has_logist = any(tok in s_low for tok in logist_tokens)
    has_surch = any(tok in s_low for tok in surch_tokens)
    if has_logist and has_surch:
        return True
    norm = normalize_sheetname_for_match(sheet_name)
    patterns = [
        normalize_sheetname_for_match("Logistik_Zuschlaege_sonst__Info"),
        normalize_sheetname_for_match("Logistics_Surcharges_Misc"),
        normalize_sheetname_for_match("Logistics___Surcharges___Misc_")
    ]
    for p in patterns:
        if difflib.get_close_matches(norm, [p], n=1, cutoff=0.6):
            return True
    return False

def parse_dim_string(s):
    if not s:
        return (None, None, None)
    s = str(s)
    m = re.search(r'(\d{1,6})\s*[x×]\s*(\d{1,6})\s*[x×]\s*(\d{1,6})', s, flags=re.I)
    if m:
        try:
            return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            return (None, None, None)
    nums = re.findall(r'(\d{1,6})', s)
    if len(nums) >= 3:
        try:
            return (int(nums[0]), int(nums[1]), int(nums[2]))
        except Exception:
            return (None, None, None)
    return (None, None, None)

# ----------------- CACHING -----------------
@st.cache_data(ttl=3600)
def load_workbook_values(file_bytes, filename):
    return openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

@st.cache_data(ttl=3600)
def load_partspec_preview_data(file_bytes, filename):
    wb = load_workbook_values(file_bytes, filename)
    sh = wb.active
    X = count_nonempty_columns_after_B(sh)
    suffixes = [
        re.sub(r'^\s*\d+\s*-\s*', '', str(sh.cell(row=19, column=c).value or ""), flags=re.I).strip()
        for c in range(3, 3 + X)
    ]
    wb.close()
    return X, suffixes

@st.cache_data(ttl=3600)
def load_plant_entries_data(file_bytes, filename):
    wb = load_workbook_values(file_bytes, filename)
    plant_sheet = next((sh for sh in wb.worksheets if "plant_upate" in sh.title.strip().lower()), None)
    if not plant_sheet:
        wb.close()
        return None
    entries = [
        (str(c).strip() if c else "", str(n).strip() if n else "")
        for c, n in plant_sheet.iter_rows(min_row=2, max_col=2, values_only=True)
        if c or n
    ]
    wb.close()
    return entries

@st.cache_data(ttl=3600, max_entries=1000)
def plant_name_matches_suffix_cached(plant_name, suffix, threshold=0.6):
    pn, sn = normalize_name(plant_name), normalize_name(suffix)
    if not pn or not sn:
        return False
    if pn == sn:
        return True
    p_tokens, s_tokens = set(pn.split()), set(sn.split())
    if p_tokens.issubset(s_tokens) and p_tokens:
        return True
    overlap_ratio = len(p_tokens & s_tokens) / len(p_tokens) if p_tokens else 0
    if overlap_ratio >= threshold:
        return True
    return bool(
        difflib.get_close_matches(pn, [sn], n=1, cutoff=0.7) or
        difflib.get_close_matches(sn, [pn], n=1, cutoff=0.7)
    )

# ----------------- SUPPLIER MATCHING -----------------
abbreviation_map = {
    "ips": "IPS ( INJECTION PLASTIQUES SYSTEMES)",
    "ips packaging": "IPS PACKAGING & AUTOMATION",
}

def find_best_supplier(norm_name, supplier_keys):
    for abbr, full in abbreviation_map.items():
        if abbr in norm_name:
            return normalize_name(full)
    name_set = set(norm_name.split())
    best_match = max(
        ((key, len(name_set & set(key.split()))) for key in supplier_keys),
        key=lambda x: x[1],
        default=(None, 0)
    )
    if best_match[1] > 0:
        return best_match[0]
    candidates = difflib.get_close_matches(norm_name, supplier_keys, n=1, cutoff=0.6)
    return candidates[0] if candidates else None

# ----------------- FILE UPLOAD -----------------
uploaded_files = st.file_uploader(
    "Upload: PartSpec_Overview, TCO, supplier_list, supplier_info (optional), and supplier quote files",
    type=["xlsx", "xlsm"], accept_multiple_files=True
)
num_files = len(uploaded_files)
st.write(f"📄 Files uploaded: **{num_files}**")

if num_files == 0:
    st.warning("⚠️ Please upload required files.")
    st.stop()

part_spec_file = template_file = supplier_list_file = supplier_info_file = None
supplier_quote_files = []
file_bytes_dict = {}

for f in uploaded_files:
    f.seek(0)
    file_bytes = f.read()
    file_bytes_dict[f.name] = file_bytes

    fname = f.name.lower()
    if "partspec_overview" in fname:
        part_spec_file = f
    elif REQUIRED_TCO_NAME_FRAGMENT in fname:
        template_file = f
    elif "supplier_list" in fname:
        supplier_list_file = f
    elif "supplier_info" in fname or "suppliers_info" in fname:
        supplier_info_file = f
    elif "_quote" in fname:
        supplier_quote_files.append(f)

if not part_spec_file or not template_file:
    st.warning("⚠️ Upload PartSpec_Overview and TCO file.")
    st.stop()

# ----------------- PREVIEW: Plant Matching -----------------
X_preview, suffixes = load_partspec_preview_data(file_bytes_dict[part_spec_file.name], part_spec_file.name)
if X_preview == 0:
    st.warning("⚠️ No data in PartSpec after column B.")
    st.stop()

plant_entries = load_plant_entries_data(file_bytes_dict[template_file.name], template_file.name)
if not plant_entries:
    st.error("❌ Could not find 'plant_upate' sheet.")
    st.stop()

auto_assign, ambiguous, no_match = {}, {}, []
for i, s in enumerate(suffixes):
    if not s.strip():
        no_match.append(i)
        continue
    matches = [(c, n) for c, n in plant_entries if plant_name_matches_suffix_cached(n, s, 0.6)]
    if not matches:
        closest = difflib.get_close_matches(s, [n for _, n in plant_entries], n=5, cutoff=0.6)
        matches = [(c, n) for c, n in plant_entries if n in closest]
    if len(matches) == 0:
        no_match.append(i)
    elif len(matches) == 1:
        auto_assign[i] = matches[0][0]
    else:
        ambiguous[i] = matches

# ----------------- INPUT FORM -----------------
with st.form("input_form"):
    st.subheader("✍️ Enter Metadata")
    status_input = st.text_input("Status")
    name_operator = st.text_input("Operator Name")
    datum = st.text_input("Datum (YYYY-MM-DD)")
    Currency = st.selectbox("Currency", ["EUR", "THB", "MYR", "USD", "CNY", "INR", "UND", "ZAR"])
    segment = st.selectbox("Segment", ["ES", "IS"])
    region = st.selectbox("Region", ["Asia", "AM", "EU"])
    carline = st.selectbox("Carline", [
        "G2Y", "Audi CBEV SOP+1", "BMW G1x", "BMW G2y Europa", "BMW G2y China",
        "BMW G2y Mex", "BMW G2y Mex + CN", "BMW G2y LCI", "BR223 I-Tafel",
        "Porsche J1 BMCe", "Porsche J1 Batterie", "Jaguar_X590", "L462", "L560",
        "M161", "PO 623", "PO 416 PA Miko a.V.", "C257 TVKL", "W213", "M156 MY19",
        "N293", "BR223 eFSR", "R232", "Porsche J1 BMCe 335t", "Porsche J1 Batterie 335t",
        "L460", "L461", "L460_L461", "MFA2", "BMW G70", "AMG HPB 150", "M182",
        "M189", "Tesla Model Y", "PO633", "VW-310", "BMW GEN6_AC_ChargeHarness",
        "AMG BR59x", "M183", "BMW G70 LCI", "BMW i 20 mulit rail",
        "JaguarX90x-500 Panthera", "L460/461 MY28 IP& CFC"
    ])

    st.markdown("### Plant Matches Preview (Column T)")
    for idx, s in enumerate(suffixes):
        st.write(f"- Col #{idx+1}: **{s or '(empty)'}**")

    if auto_assign:
        st.write("### Auto-assigned:")
        for idx, code in auto_assign.items():
            st.write(f"- Col {idx+1}: `{code}`")

    if no_match:
        st.warning("No match: " + ", ".join(str(i+1) for i in no_match))

    selection_map = {}
    for idx, opts in ambiguous.items():
        choices = [f"{c} - {n}" for c, n in opts]
        selection_map[idx] = st.selectbox(
            f"Select plant for Column {idx+1} — '{suffixes[idx]}'",
            choices,
            key=f"plant_{idx}"
        ).split(" - ", 1)[0].strip()

    # NOTE: Variant Chain UI inputs removed — column U will be auto-filled from Master Data mapping
    submit = st.form_submit_button("Process Excel")

# ----------------- PROCESSING -----------------
if submit:
    if not all([status_input, name_operator, datum, segment, region, carline]):
        st.warning("⚠️ Please fill in all fields.")
        st.stop()
    if not supplier_list_file or not supplier_quote_files:
        st.error("❌ Missing supplier_list or quote files.")
        st.stop()

    # Load partspec
    wb_part = load_workbook_values(file_bytes_dict[part_spec_file.name], part_spec_file.name)
    sh_part = wb_part.active
    X = count_nonempty_columns_after_B(sh_part)
    col_indices = list(range(3, 3 + X))

    # Prefetch rows
    plmids = [str(sh_part.cell(row=6, column=c).value or "").split("/")[0].strip() for c in col_indices]
    vals_Z = [sh_part.cell(row=5, column=c).value for c in col_indices]
    vals_AA = [sh_part.cell(row=31, column=c).value for c in col_indices]
    vals_AB = [sh_part.cell(row=56, column=c).value for c in col_indices]
    af_vals = [sh_part.cell(row=26, column=c).value or 0 for c in col_indices]
    # NEW: partspec row 19 plant names (used for Master Data AI matching)
    partspec_row19 = [sh_part.cell(row=19, column=c).value for c in col_indices]

    # Safety stock AV (row69 first non-empty after B)
    safety_value = None
    for c in col_indices:
        v = sh_part.cell(row=69, column=c).value
        if v not in (None, ""):
            try:
                safety_value = float(v)
            except Exception:
                m = re.search(r'[-+]?\d*\.?\d+', str(v))
                safety_value = float(m.group(0)) if m else 0.0
            break
    if safety_value is None:
        safety_value = 0.0

    # BA and dims (row70 per-column) and BI detection + BP/BQ parse base values
    BA_target_values = [None] * X
    dims_values = [(None, None, None) for _ in range(X)]
    bi_values = [None] * X
    bp_by_part = [None] * X
    bq_by_part = [None] * X

    for idx, c in enumerate(col_indices):
        v = sh_part.cell(row=70, column=c).value
        if v not in (None, ""):
            s = str(v)
            # BI: cardboard -> karton
            if re.search(r'cardboard', s, flags=re.I):
                bi_values[idx] = "karton"
            else:
                bi_values[idx] = None
            # dims from row70
            dims_values[idx] = parse_dim_string(s)
            # BA target number before 'pcs'
            mpcs = re.search(r'(\d+)\s*(?=pcs)', s, flags=re.I)
            if not mpcs:
                mpcs = re.search(r'(\d+)\s*(?=pcs\b)', s, flags=re.I)
            if not mpcs:
                m2 = re.search(r'(\d+)', s)
                if m2:
                    mpcs = m2
            if mpcs:
                try:
                    BA_target_values[idx] = int(mpcs.group(1))
                except Exception:
                    try:
                        BA_target_values[idx] = int(float(mpcs.group(1)))
                    except Exception:
                        BA_target_values[idx] = None
            # --- Parse BP/BQ once (used for ALL blocks) ---
            s_lower = s.lower()
            m_bp = BOX_PER_LAYER_RE.search(s_lower)
            if m_bp:
                try:
                    bp_by_part[idx] = int(m_bp.group(1))
                except Exception:
                    bp_by_part[idx] = None
            m_bq = LAYERS_RE.search(s_lower)
            if m_bq:
                try:
                    bq_by_part[idx] = int(m_bq.group(1))
                except Exception:
                    bq_by_part[idx] = None
        else:
            dims_values[idx] = (None, None, None)
            bi_values[idx] = None
            BA_target_values[idx] = None
            bp_by_part[idx] = None
            bq_by_part[idx] = None

    # Supplier list
    wb_sup = load_workbook_values(file_bytes_dict[supplier_list_file.name], supplier_list_file.name)
    sh_sup = wb_sup.active
    supplier_dict = {
        normalize_name(name): str(num).strip()
        for num, name in sh_sup.iter_rows(min_row=2, max_col=2, values_only=True)
        if num and name
    }
    supplier_keys = list(supplier_dict.keys())
    wb_sup.close()

    # Suppliers_info optional
    suppliers_lookup = {}
    suppliers_lookup_E = {}
    suppliers_lookup_Q = {}
    suppliers_lookup_R = {}
    if supplier_info_file:
        try:
            wb_supplier_info = load_workbook_values(file_bytes_dict[supplier_info_file.name], supplier_info_file.name)
            s_info_sheet = wb_supplier_info.active
            for row in s_info_sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 18 and row[0]:
                    code_A = str(row[0]).strip()
                    suppliers_lookup[code_A] = row[5] if len(row) > 5 else None
                    suppliers_lookup_E[code_A] = row[4] if len(row) > 4 else None
                    suppliers_lookup_Q[code_A] = row[16] if len(row) > 16 else None
                    suppliers_lookup_R[code_A] = row[17] if len(row) > 17 else None
            wb_supplier_info.close()
            st.info(f"✅ Loaded Suppliers_info with {len(suppliers_lookup)} entries")
        except Exception as e:
            st.warning(f"⚠️ Failed to process Suppliers_info file: {e}")
    else:
        st.info("ℹ️ No Suppliers_info file provided - skipping column AQ-AU processing")

    # Load template workbook
    template_bytes = BytesIO(file_bytes_dict[template_file.name])
    wb = openpyxl.load_workbook(template_bytes, keep_vba=True)
    if len(wb.worksheets) < 4:
        st.error("❌ TCO workbook must have at least 4 sheets.")
        st.stop()
    sheet_template = wb.worksheets[2]
    sheet_master = wb.worksheets[3]

    try:
        sheet_template.freeze_panes = None
    except Exception:
        pass

    plant_sheet = next((sh for sh in wb.worksheets if "plant_upate" in sh.title.strip().lower()), None)
    if not plant_sheet:
        st.error("❌ 'plant_upate' sheet not found.")
        st.stop()

    codes_for_cols = [auto_assign.get(i) or selection_map.get(i) for i in range(X)]

    # ----------------- STYLE CACHES & apply_row_style up to BX -----------------
    fill_cache = {}
    border_cache = {}

    def random_unique_color():
        for _ in range(100):
            r, g, b = [random.randint(150, 255) for _ in range(3)]
            color = f"{r:02X}{g:02X}{b:02X}"
            if color not in used_colors:
                used_colors.add(color)
                return color
        return "CCCCCC"

    def get_fill(color_hex):
        if color_hex not in fill_cache:
            fill_cache[color_hex] = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        return fill_cache[color_hex]

    def get_border(left=False, right=False):
        key = (left, right)
        if key not in border_cache:
            border_cache[key] = Border(
                left=thin if left else None,
                right=thin if right else None,
                top=thin,
                bottom=thin
            )
        return border_cache[key]

    def apply_row_style(sheet, row_num, fill_color=None):
        fill_color = fill_color or random_unique_color()
        fill_obj = get_fill(fill_color)
        max_col = STYLE_MAX_COL_NUM  # extend styling up to BX
        cell_fn = sheet.cell
        bold_font = Font(bold=True)
        left_border = get_border(left=True, right=False)
        right_border = get_border(left=False, right=True)
        mid_border = get_border(left=False, right=False)
        for col in range(1, max_col + 1):
            cell = cell_fn(row=row_num, column=col)
            try:
                if isinstance(cell, openpyxl.cell.MergedCell):
                    continue
            except Exception:
                pass
            cell.fill = fill_obj
            if col == 1:
                cell.border = left_border
            elif col == max_col:
                cell.border = right_border
            else:
                cell.border = mid_border
            if col == 19:  # Column S bold
                cell.font = bold_font

    # ----------------- PRESERVE TEMPLATE ROWS (10 & 11) FORMULAS -----------------
    formula_cells_template_header = {}
    formula_cells_template_data = {}
    # Base columns to copy formulas from (existing behavior) + add BR..BX (70..76)
    base_cols = (
        [chr(i) for i in range(ord('L'), ord('Q') + 1)] +  # L-Q
        ['V', 'W'] +
        [get_column_letter(i) for i in range(33, 38)] +    # AG-AL (33..37)
        [get_column_letter(i) for i in range(70, 77)]      # BR-BX (70..76)
    )
    for c in base_cols:
        try:
            cell_h = sheet_template[f"{c}{TEMPLATE_HEADER_ROW}"]
            if isinstance(cell_h.value, str) and cell_h.value.startswith('='):
                formula_cells_template_header[c] = cell_h.value
        except Exception:
            pass
        try:
            cell_d = sheet_template[f"{c}{TEMPLATE_DATA_ROW}"]
            if isinstance(cell_d.value, str) and cell_d.value.startswith('='):
                formula_cells_template_data[c] = cell_d.value
        except Exception:
            pass

    # ensure S at new header & style
    sheet_template[f"S{HEADER_ROW}"].value = "Target"
    apply_row_style(sheet_template, HEADER_ROW)
    sheet_template[f"B{DATA_ROW}"].value = None

    # ----------------- CLEAR NEW HEADER/DATA CELLS (NOT rows 10 & 11) -----------------
    metadata_cols_to_clear = ['G', 'H', 'J', 'T'] + [get_column_letter(i) for i in range(27, 29)]
    for col in metadata_cols_to_clear:
        sheet_template[f"{col}{HEADER_ROW}"].value = None
    new_cols_to_clear = [get_column_letter(i) for i in range(43, 48)]
    for col in new_cols_to_clear:
        sheet_template[f"{col}{HEADER_ROW}"].value = None

    # Clear BJ,BK,BL and BD/BH in new header/data rows only
    BJ_col_letter = get_column_letter(62)
    BK_col_letter = get_column_letter(63)
    BL_col_letter = get_column_letter(64)
    for row_to_clear in (HEADER_ROW, DATA_ROW):
        sheet_template[f"{BJ_col_letter}{row_to_clear}"].value = None
        sheet_template[f"{BK_col_letter}{row_to_clear}"].value = None
        sheet_template[f"{BL_col_letter}{row_to_clear}"].value = None

    BD_col_letter = get_column_letter(BD_col)
    BH_col_letter = get_column_letter(BH_col)
    for row_to_clear in (HEADER_ROW, DATA_ROW):
        sheet_template[f"{BD_col_letter}{row_to_clear}"].value = None
        sheet_template[f"{BH_col_letter}{row_to_clear}"].value = None

    # ----------------- BATCH FILL METADATA -----------------
    num_blocks = len(supplier_quote_files) + 1
    rows_to_write = [DATA_ROW + block * (X + 1) + i for block in range(num_blocks) for i in range(X)]
    metadata = {
        "C": status_input,
        "D": name_operator,
        "E": datum,
        "G": Currency,
        "I": segment,
        "J": region,
        "K": carline,
        "H": "1"
    }
    cell_fn = sheet_template.cell
    for col, val in metadata.items():
        col_index = column_index_from_string(col)
        for row in rows_to_write:
            cell_fn(row=row, column=col_index).value = val

    # ----------------- SCENARIO INDEX (F) -----------------
    for block in range(num_blocks):
        base = DATA_ROW + block * (X + 1)
        for i in range(X):
            r = base + i
            c = cell_fn(row=r, column=6)
            try:
                if not isinstance(c, openpyxl.cell.MergedCell):
                    c.value = str(block)
            except Exception:
                c.value = str(block)

    # ----------------- COPY FORMULAS (incl. BR..BX) -----------------
    for block in range(num_blocks):
        base = DATA_ROW + block * (X + 1)
        for i in range(X):
            r = base + i
            for col, formula in formula_cells_template_header.items():
                try:
                    sheet_template[f"{col}{r}"].value = shift_formula_rows(formula, r - TEMPLATE_HEADER_ROW)
                except Exception:
                    pass
            for col, formula in formula_cells_template_data.items():
                try:
                    sheet_template[f"{col}{r}"].value = shift_formula_rows(formula, r - TEMPLATE_DATA_ROW)
                except Exception:
                    pass

    # ----------------- SUPPLIER DATA (preload workbooks) -----------------
    R_col, S_col = 18, 19
    current_row = HEADER_ROW + (X + 1)
    supplier_plmid_data = {}
    supplier_logistics_map = {}

    supplier_wb_cache = {}
    for file in supplier_quote_files:
        try:
            file_bytes = file_bytes_dict[file.name]
            wb_sup_quote = load_workbook_values(file_bytes, file.name)
            supplier_wb_cache[file.name] = wb_sup_quote
        except Exception as e:
            st.warning(f"Failed to read {file.name}: {e}")
            supplier_wb_cache[file.name] = None

    for file in supplier_quote_files:
        wb_sup_quote = supplier_wb_cache.get(file.name)
        if not wb_sup_quote:
            supplier_plmid_data[file.name] = {}
            supplier_logistics_map[file.name] = {}
            continue

        plmid_dict = {}
        logi_map = {}
        assigned_indices = set()
        matched_sheets_no_digit = []

        try:
            for sh_name in wb_sup_quote.sheetnames:
                # PLMID logic
                m_num = re.search(r'(\d+)', sh_name)
                if m_num:
                    key = m_num.group(1)
                    try:
                        sh = wb_sup_quote[sh_name]
                        plmid_dict[key] = {
                            "D49": sh.cell(row=49, column=4).value,
                            "D51": sh.cell(row=51, column=4).value
                        }
                    except Exception:
                        plmid_dict[key] = {"D49": None, "D51": None}

                # detect logistics sheet robustly
                if is_logistics_sheet(sh_name):
                    nums = re.findall(r'(\d+)', sh_name)
                    try:
                        sh = wb_sup_quote[sh_name]
                        d33 = sh.cell(row=33, column=4).value
                        d32_raw = sh.cell(row=32, column=4).value
                        dims = parse_dim_string(d32_raw)
                    except Exception:
                        d33 = None
                        dims = (None, None, None)

                    if nums:
                        idx_num = int(nums[-1]) - 1
                        if 0 <= idx_num < X and idx_num not in assigned_indices:
                            logi_map[idx_num] = {"D33": d33, "dims": dims}
                            assigned_indices.add(idx_num)
                            continue
                        matched_sheets_no_digit.append((sh_name, d33, dims))
                    else:
                        matched_sheets_no_digit.append((sh_name, d33, dims))

            # assign unmatched logistics sheets to remaining part indices
            remaining_indices = [i for i in range(X) if i not in assigned_indices]
            for (sh_name, d33, dims), part_idx in zip(matched_sheets_no_digit, remaining_indices):
                logi_map[part_idx] = {"D33": d33, "dims": dims}
                assigned_indices.add(part_idx)

            supplier_plmid_data[file.name] = plmid_dict
            supplier_logistics_map[file.name] = logi_map

        except Exception as e:
            st.warning(f"Failed to process sheets in {file.name}: {e}")
            supplier_plmid_data[file.name] = {}
            supplier_logistics_map[file.name] = {}

        # write supplier header + R fill as before
        raw_name = extract_supplier_name_from_filename(file.name)
        norm_name = normalize_name(raw_name)
        supplier_number = supplier_dict.get(norm_name, "N/A")
        if supplier_number == "N/A":
            match_key = find_best_supplier(norm_name, supplier_keys)
            supplier_number = supplier_dict.get(match_key, "N/A")

        sheet_template.cell(row=current_row, column=S_col).value = raw_name
        for i in range(1, X + 1):
            sheet_template.cell(row=current_row + i, column=R_col).value = supplier_number
        apply_row_style(sheet_template, current_row)
        current_row += X + 1

    # close supplier workbooks
    for wbq in supplier_wb_cache.values():
        try:
            if wbq:
                wbq.close()
        except Exception:
            pass

    # ----------------- PARTSPEC DATA & WRITE -----------------
    T_col, X_col, Z_col, AA_col, AB_col = 20, 24, 26, 27, 28
    AL_col, AN_col, AO_col, AF_col = 38, 40, 41, 32
    AV_col, BA_col, BI_col, BJ_col, BK_col, BL_col, BM_col = 48, 53, 61, 62, 63, 64, 65
    BN_col = BN_COL_NUM  # 66

    for block, file in enumerate([None] + supplier_quote_files):
        base = DATA_ROW + block * (X + 1)
        for i in range(X):
            r = base + i
            sheet_template.cell(row=r, column=T_col).value = codes_for_cols[i]
            sheet_template.cell(row=r, column=X_col).value = plmids[i]
            sheet_template.cell(row=r, column=Z_col).value = vals_Z[i]
            sheet_template.cell(row=r, column=AA_col).value = vals_AA[i]
            try:
                abv = float(vals_AB[i]) if vals_AB[i] not in (None, "") else None
                sheet_template.cell(row=r, column=AB_col).value = abv
            except (ValueError, TypeError):
                sheet_template.cell(row=r, column=AB_col).value = None

            # ----------------- BD, BE, BH, BO (packaging flags) -----------------
            try:
                part_col = col_indices[i]
                raw70 = sh_part.cell(row=70, column=part_col).value
                s70 = str(raw70).lower() if raw70 not in (None, "") else ""
            except Exception:
                s70 = ""

            bd_val = None
            bo_val = None
            # check for ow-pallet / ow-pal
            if 'ow-pallet' in s70 or 'ow-pal' in s70:
                bd_val = 'd'
                bo_val = 'pallet one way'
            # check for KLT (foldable KLT or KLT)
            elif 'klt' in s70:
                bd_val = 'r'
                bo_val = 'returnable pallet'

            # write BD (56), BE (57), BO (67), BH (60)
            try:
                sheet_template.cell(row=r, column=BD_col).value = bd_val
            except Exception:
                pass
            try:
                sheet_template.cell(row=r, column=BH_col).value = bd_val
            except Exception:
                pass
            try:
                sheet_template.cell(row=r, column=BE_col).value = 'homo'
            except Exception:
                pass
            try:
                sheet_template.cell(row=r, column=BO_col).value = bo_val
            except Exception:
                pass

            # ----------------- BP & BQ: use parsed values for ALL blocks -----------------
            bp_val = bp_by_part[i]
            bq_val = bq_by_part[i]
            try:
                if bp_val is not None:
                    sheet_template.cell(row=r, column=BP_col).value = bp_val
            except Exception:
                pass
            try:
                if bq_val is not None:
                    sheet_template.cell(row=r, column=BQ_col).value = bq_val
            except Exception:
                pass

            # Supplier PLMID-related cells
            if file and plmids[i]:
                suffix = str(plmids[i])[1:] if len(str(plmids[i])) > 1 else ""
                data = supplier_plmid_data.get(file.name, {}).get(suffix, {})
                sheet_template.cell(row=r, column=AL_col).value = data.get("D51")
                sheet_template.cell(row=r, column=AN_col).value = data.get("D49")
                sheet_template.cell(row=r, column=AO_col).value = data.get("D51")
            else:
                for col in (AL_col, AN_col, AO_col):
                    sheet_template.cell(row=r, column=col).value = None

            # AV: safety stock (float with one decimal)
            cell_av = sheet_template.cell(row=r, column=AV_col)
            try:
                cell_av.value = float(safety_value)
            except Exception:
                cell_av.value = float(0.0)
            cell_av.number_format = "0.0"

            # BA and BM
            cell_ba = sheet_template.cell(row=r, column=BA_col)
            cell_bm = sheet_template.cell(row=r, column=BM_col)
            if block == 0:
                val = BA_target_values[i]
                cell_ba.value = val if val is not None else None
                cell_bm.value = val if val is not None else None
            else:
                logi_map = supplier_logistics_map.get(file.name, {})
                entry = logi_map.get(i)
                if entry is None:
                    entry = logi_map.get(str(plmids[i])[1:] if len(str(plmids[i])) > 1 else "")
                if entry is None:
                    entry = logi_map.get(str(i + 1))
                if isinstance(entry, dict):
                    cell_ba.value = entry.get("D33")
                    cell_bm.value = entry.get("D33")
                else:
                    cell_ba.value = entry
                    cell_bm.value = entry

            # BI (karton) — now for ALL blocks
            cell_bi = sheet_template.cell(row=r, column=BI_col)
            cell_bi.value = bi_values[i]

            # BJ/BK/BL dims
            cell_bj = sheet_template.cell(row=r, column=BJ_col)
            cell_bk = sheet_template.cell(row=r, column=BK_col)
            cell_bl = sheet_template.cell(row=r, column=BL_col)

            if block == 0:
                d1, d2, d3 = dims_values[i]
                cell_bj.value = d1
                cell_bk.value = d2
                cell_bl.value = d3
            else:
                logi_map = supplier_logistics_map.get(file.name, {})
                entry = logi_map.get(i)
                if entry is None:
                    entry = logi_map.get(str(plmids[i])[1:] if len(str(plmids[i])) > 1 else "")
                if entry is None:
                    entry = logi_map.get(str(i + 1))
                if isinstance(entry, dict):
                    d1, d2, d3 = entry.get("dims", (None, None, None))
                    cell_bj.value = d1
                    cell_bk.value = d2
                    cell_bl.value = d3
                else:
                    cell_bj.value = None
                    cell_bk.value = None
                    cell_bl.value = None

    # ----------------- COLUMN U: fill from Master Data based on first-two-chars of T and plant name match (AI) -----------------
    # Build master data rows: AF (32) -> AG (33) and AI (35)
    master_sheet = None
    # Robust search for a sheet whose name contains both 'master' and 'data'
    for sh in wb.worksheets:
        t = sh.title.strip().lower()
        if "master" in t and "data" in t:
            master_sheet = sh
            break
    # fallback to sheet_master if not found above
    if master_sheet is None:
        try:
            if sheet_master and isinstance(sheet_master, openpyxl.worksheet.worksheet.Worksheet):
                master_sheet = sheet_master
        except Exception:
            master_sheet = None

    master_by_af = {}  # AF_key -> list of dicts { 'ag': val, 'ai': ai_norm, 'af_raw': af_raw }
    master_rows = []    # list of entries for fallback fuzzy / AI-only matching
    if master_sheet:
        for row in master_sheet.iter_rows(min_row=2, min_col=32, max_col=35, values_only=True):
            af_raw, ag_val, ah_val, ai_val = None, None, None, None
            # row returns AF, AG, AH, AI because max_col=35; safe destructure:
            # But some files could have fewer columns; handle robustly:
            # In our iter_rows we requested 32..35, values_only returns tuple length 4
            try:
                af_raw = row[0]
                ag_val = row[1] if len(row) > 1 else None
                # row[2] is AH (ignored), row[3] is AI
                ai_val = row[3] if len(row) > 3 else None
            except Exception:
                continue
            if af_raw is None and ai_val is None:
                continue
            af_key = str(af_raw).strip().upper() if af_raw is not None else ""
            ai_norm = normalize_name(ai_val) if ai_val is not None else ""
            entry = {'af_raw': af_key, 'ag': ag_val, 'ai_norm': ai_norm}
            master_rows.append(entry)
            if af_key:
                master_by_af.setdefault(af_key, []).append(entry)

    # Fill U using combined AF key and AI plant-name match (partspec row19)
    for block in range(num_blocks):
        base = DATA_ROW + block * (X + 1)
        for i in range(X):
            r = base + i
            t_val = sheet_template.cell(row=r, column=T_col).value
            if not t_val:
                continue
            t_str = str(t_val).strip().upper()
            # extract first two characters as requested (if only one char present, use it)
            key = t_str[:2] if len(t_str) >= 2 else t_str[:1]
            # plant name from partspec row19 for that column
            p19 = partspec_row19[i]
            partspec_plant_norm = normalize_name(p19) if p19 else ""
            chosen_ag = None

            # 1) exact AF key matches: prefer the one whose AI matches partspec plant name
            candidates = master_by_af.get(key, [])
            if candidates:
                if partspec_plant_norm:
                    # try to find candidate where ai_norm equals or contains partspec_plant_norm
                    matched = None
                    for cand in candidates:
                        cand_ai = cand.get('ai_norm') or ""
                        if cand_ai and (cand_ai == partspec_plant_norm or partspec_plant_norm in cand_ai or cand_ai in partspec_plant_norm):
                            matched = cand
                            break
                    if matched:
                        chosen_ag = matched.get('ag')
                    else:
                        # no AI match among AF-candidates -> take first AF candidate
                        chosen_ag = candidates[0].get('ag')
                else:
                    chosen_ag = candidates[0].get('ag')

            # 2) if no AF candidate matched, try fuzzy AF match
            if chosen_ag is None and master_by_af:
                af_keys = list(master_by_af.keys())
                fuzzy = difflib.get_close_matches(key, af_keys, n=1, cutoff=0.8)
                if fuzzy:
                    fcand = master_by_af.get(fuzzy[0], [])
                    if fcand:
                        # if plant name available, prefer matching AI among fuzzy candidates
                        if partspec_plant_norm:
                            matched = None
                            for cand in fcand:
                                cand_ai = cand.get('ai_norm') or ""
                                if cand_ai and (cand_ai == partspec_plant_norm or partspec_plant_norm in cand_ai or cand_ai in partspec_plant_norm):
                                    matched = cand
                                    break
                            if matched:
                                chosen_ag = matched.get('ag')
                            else:
                                chosen_ag = fcand[0].get('ag')
                        else:
                            chosen_ag = fcand[0].get('ag')

            # 3) if still not found, try matching by AI (plant name) only
            if chosen_ag is None and partspec_plant_norm:
                for cand in master_rows:
                    cand_ai = cand.get('ai_norm') or ""
                    if cand_ai and (cand_ai == partspec_plant_norm or partspec_plant_norm in cand_ai or cand_ai in partspec_plant_norm):
                        chosen_ag = cand.get('ag')
                        break

            # 4) fallback: try exact key in master_rows by af_raw ignoring case/punct
            if chosen_ag is None and master_rows:
                key_norm_plain = re.sub(r'[^0-9A-Z]', '', key.upper())
                for cand in master_rows:
                    afcand = cand.get('af_raw') or ""
                    afcand_norm = re.sub(r'[^0-9A-Z]', '', str(afcand).upper())
                    if afcand_norm == key_norm_plain:
                        chosen_ag = cand.get('ag')
                        break

            if chosen_ag is not None:
                try:
                    sheet_template.cell(row=r, column=U_col).value = chosen_ag
                except Exception:
                    pass

    # ----------------- COLUMN AF (percentage) -----------------
    for block in range(num_blocks):
        base = DATA_ROW + block * (X + 1)
        for i in range(X):
            r = base + i
            cell = sheet_template.cell(row=r, column=AF_col)
            # Robustly convert source AF value into a numeric literal for the formula
            raw = af_vals[i]
            num = 0.0
            try:
                if raw is None or raw == "":
                    num = 0.0
                elif isinstance(raw, str):
                    s = raw.strip()
                    if '%' in s:
                        s2 = s.replace('%', '').replace(',', '.')
                        num = float(s2) / 100.0
                    else:
                        s2 = s.replace(',', '.')
                        num = float(s2)
                else:
                    num = float(raw)
            except Exception:
                num = 0.0
            try:
                cell.value = f"={num}/M{r}"
                cell.number_format = "0%"
            except Exception:
                pass

    # ----------------- AQ and AR-AU processing (unchanged) -----------------
    if suppliers_lookup:
        AQ_col = 43
        sheet_template.cell(row=DATA_ROW, column=AQ_col).value = None
        for block in range(1, num_blocks):
            base = DATA_ROW + block * (X + 1)
            for i in range(X):
                r = base + i
                col_T_value = sheet_template.cell(row=r, column=20).value
                col_R_value = sheet_template.cell(row=r, column=18).value
                code = ""
                if col_T_value and col_R_value:
                    code = f"{col_T_value}{col_R_value}"
                elif col_T_value:
                    code = f"{col_T_value}"
                sheet_template.cell(row=r, column=AQ_col).value = suppliers_lookup.get(code)
    else:
        st.info("⏭️ Skipping column AQ processing (no Suppliers_info file)")

    if suppliers_lookup:
        AR_col, AS_col, AT_col, AU_col = 44, 45, 46, 47
        for col in (AR_col, AS_col, AT_col, AU_col):
            sheet_template.cell(row=DATA_ROW, column=col).value = None
        for block in range(1, num_blocks):
            base = DATA_ROW + block * (X + 1)
            for i in range(X):
                r = base + i
                col_T_value = sheet_template.cell(row=r, column=20).value
                col_R_value = sheet_template.cell(row=r, column=18).value
                code = ""
                if col_T_value and col_R_value:
                    code = f"{col_T_value}{col_R_value}"
                elif col_T_value:
                    code = f"{col_T_value}"

                ar_value = 0.0
                ev = suppliers_lookup_E.get(code)
                if ev is not None:
                    try:
                        ar_value = float(ev)
                        sheet_template.cell(row=r, column=AR_col).value = ar_value
                    except (ValueError, TypeError):
                        sheet_template.cell(row=r, column=AR_col).value = 0.0
                else:
                    sheet_template.cell(row=r, column=AR_col).value = 0.0

                as_value = 0.0
                qv = suppliers_lookup_Q.get(code)
                if qv is not None:
                    try:
                        as_value = float(qv)
                        sheet_template.cell(row=r, column=AS_col).value = as_value
                    except (ValueError, TypeError):
                        sheet_template.cell(row=r, column=AS_col).value = 0.0
                else:
                    sheet_template.cell(row=r, column=AS_col).value = 0.0

                at_value = 0.0
                rv = suppliers_lookup_R.get(code)
                if rv is not None:
                    try:
                        at_value = float(rv)
                        sheet_template.cell(row=r, column=AT_col).value = at_value
                    except (ValueError, TypeError):
                        sheet_template.cell(row=r, column=AT_col).value = 0.0
                else:
                    sheet_template.cell(row=r, column=AT_col).value = 0.0

                sheet_template.cell(row=r, column=AU_col).value = as_value + at_value
    else:
        st.info("⏭️ Skipping columns AR-AU processing (no Suppliers_info file)")

    # ----------------- BN: copy data-row formula; keep supplier headers empty ---------------
    try:
        raw_bn_template_header = sheet_template[f"{get_column_letter(BN_col)}{TEMPLATE_HEADER_ROW}"].value
        raw_bn_template_data = sheet_template[f"{get_column_letter(BN_col)}{TEMPLATE_DATA_ROW}"].value
        bn_header_val = raw_bn_template_header if isinstance(raw_bn_template_header, str) and raw_bn_template_header.startswith('=') else None
        bn_data_val = raw_bn_template_data if isinstance(raw_bn_template_data, str) and raw_bn_template_data.startswith('=') else None
    except Exception:
        bn_header_val = None
        bn_data_val = None

    # Copy BN data-row formula into data rows of every block (shifted)
    if bn_data_val:
        for block in range(num_blocks):
            base = DATA_ROW + block * (X + 1)
            for i in range(X):
                r = base + i
                try:
                    shifted = shift_formula_rows(bn_data_val, r - TEMPLATE_DATA_ROW)
                    sheet_template.cell(row=r, column=BN_col).value = shifted
                except Exception:
                    pass

    # Copy BN header formula only into Target header (block 0) but leave supplier headers empty
    if bn_header_val:
        header_row_target = HEADER_ROW
        try:
            sheet_template.cell(row=header_row_target, column=BN_col).value = shift_formula_rows(bn_header_val, HEADER_ROW - TEMPLATE_HEADER_ROW)
        except Exception:
            pass
    for block in range(1, num_blocks):
        header_row = HEADER_ROW + block * (X + 1)
        sheet_template.cell(row=header_row, column=BN_col).value = None

    # Re-apply styling for header rows across A..BX
    for block in range(num_blocks):
        header_row = HEADER_ROW + block * (X + 1)
        apply_row_style(sheet_template, header_row)

    # ----------------- FINALIZE -----------------
    wb_part.close()
    output = BytesIO()
    try:
        wb.save(output)
    except Exception as e:
        st.error(f"Failed to save workbook: {e}")
        wb.close()
        st.stop()
    output.seek(0)
    wb.close()

    st.success("✅ Processing complete — formulas (incl. BR→BX) copied; BP/BQ applied to all blocks; column U filled from Master Data (AF+AI logic); A..BX styling applied; rows 10–11 preserved.")
    st.download_button(
        label="⬇️ Download Updated Excel File",
        data=output,
        file_name=f"updated_{template_file.name}",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
